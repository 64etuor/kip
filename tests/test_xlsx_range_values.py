from __future__ import annotations

import json
import zipfile
from datetime import date, datetime, time, timedelta
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from typer.testing import CliRunner

from kip.adapters.parsers.xlsx import read_xlsx_range
from kip.adapters.storage.local import LocalWorkbookReader
from kip.api import create_app
from kip.cli import app
from kip.domain.models import SearchRequest
from kip.errors import ValidationError


def _rewrite_sheet_xml(path: Path, replacements: dict[bytes, bytes]) -> None:
    _rewrite_member(path, "xl/worksheets/sheet1.xml", replacements)


def _rewrite_member(path: Path, member_name: str, replacements: dict[bytes, bytes]) -> None:
    rewritten = path.with_suffix(".rewritten.xlsx")
    with zipfile.ZipFile(path) as source, zipfile.ZipFile(
        rewritten,
        "w",
        compression=zipfile.ZIP_DEFLATED,
    ) as target:
        for member in source.infolist():
            payload = source.read(member.filename)
            if member.filename == member_name:
                for old, new in replacements.items():
                    assert old in payload
                    payload = payload.replace(old, new, 1)
            target.writestr(member, payload)
    rewritten.replace(path)


def test_local_reader_normalizes_every_supported_openpyxl_scalar(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.iso_dates = True
    sheet = workbook.active
    sheet.title = "Values"
    sheet.append(
        [
            date(2026, 8, 13),
            datetime(2026, 8, 13, 9, 30, 15),
            time(9, 30, 15),
            timedelta(days=1, hours=2, minutes=3, seconds=4, milliseconds=500),
            True,
            "#DIV/0!",
        ]
    )
    sheet["A1"].number_format = "yyyy-mm-dd"
    sheet["B1"].number_format = "yyyy-mm-dd hh:mm:ss"
    sheet["C1"].number_format = "hh:mm:ss"
    sheet["D1"].number_format = "[h]:mm:ss.000"
    sheet["F1"].data_type = "e"
    path = tmp_path / "values.xlsx"
    workbook.save(path)

    cells = LocalWorkbookReader().read(path, "Values", "A1:F1")[0]

    assert [cell["value"] for cell in cells] == [
        "2026-08-13",
        "2026-08-13T09:30:15",
        "09:30:15",
        "P1DT2H3M4.5S",
        True,
        "#DIV/0!",
    ]
    assert [cell["value_type"] for cell in cells] == [
        "date",
        "datetime",
        "time",
        "duration",
        "boolean",
        "error",
    ]
    assert cells[0]["excel_serial"] == pytest.approx(46247.0)
    assert cells[3]["excel_serial"] == pytest.approx(1.08546875)
    json.dumps(cells, allow_nan=False)


def test_formula_objects_and_date_cache_are_readable_json(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Formulas"
    sheet["A1"] = "=DATE(2026,8,13)"
    sheet["A1"].number_format = "yyyy-mm-dd"
    sheet["B1"] = ArrayFormula("B1:B2", "=ROW(B1:B2)*2")
    sheet["C1"] = DataTableFormula(
        "C1:D2",
        ca=True,
        dt2D=True,
        dtr=True,
        r1="E1",
        r2="E2",
    )
    path = tmp_path / "formulas.xlsx"
    workbook.save(path)
    _rewrite_sheet_xml(
        path,
        {b"<f>DATE(2026,8,13)</f><v></v>": b"<f>DATE(2026,8,13)</f><v>46247</v>"},
    )

    cells = LocalWorkbookReader().read(path, "Formulas", "A1:C1")[0]

    assert cells[0]["value"] == "=DATE(2026,8,13)"
    assert cells[0]["cached_value"] == "2026-08-13T00:00:00"
    assert cells[0]["cached_excel_serial"] == pytest.approx(46247.0)
    assert cells[0]["formula_kind"] == "normal"
    assert cells[1]["value"] == "=ROW(B1:B2)*2"
    assert cells[1]["formula"] == "=ROW(B1:B2)*2"
    assert cells[1]["formula_kind"] == "array"
    assert cells[1]["formula_ref"] == "B1:B2"
    assert cells[2]["value"] is None
    assert cells[2]["formula_kind"] == "data_table"
    assert cells[2]["formula_ref"] == "C1:D2"
    assert cells[2]["formula_attributes"]["dt2D"] == "1"
    assert "openpyxl" not in json.dumps(cells)


def test_range_read_returns_the_exact_requested_rectangle(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sparse"
    sheet["B2"] = "used"
    path = tmp_path / "sparse.xlsx"
    workbook.save(path)

    single = read_xlsx_range(path, "Sparse", "B2")
    empty = read_xlsx_range(path, "Sparse", "C3:D5")

    assert [[cell["coordinate"] for cell in row] for row in single["cells"]] == [["B2"]]
    assert single["cells"][0][0]["value"] == "used"
    assert [[cell["coordinate"] for cell in row] for row in empty["cells"]] == [
        ["C3", "D3"],
        ["C4", "D4"],
        ["C5", "D5"],
    ]
    assert all(cell["value"] is None for row in empty["cells"] for cell in row)


@pytest.mark.parametrize(
    "cell_range",
    [
        "C2:A1",
        "XFE1:XFE1",
        "A1048577:A1048577",
        "A1:B50001",
    ],
)
def test_range_read_rejects_reversed_out_of_bounds_or_oversized_ranges(
    tmp_path: Path,
    cell_range: str,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Bounds"
    path = tmp_path / "bounds.xlsx"
    workbook.save(path)

    with pytest.raises(ValidationError, match="range"):
        read_xlsx_range(path, "Bounds", cell_range)


def test_range_read_marks_merged_and_filtered_cells(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Layout"
    sheet["A1"] = "merged"
    sheet.merge_cells("A1:B2")
    sheet.append(["header", "value"])
    sheet.append(["visible", 1])
    sheet.append(["filtered", 2])
    sheet.auto_filter.ref = "A3:B5"
    sheet.row_dimensions[2].hidden = True
    sheet.row_dimensions[5].hidden = True
    path = tmp_path / "layout.xlsx"
    workbook.save(path)

    cells = read_xlsx_range(path, "Layout", "A1:B5")["cells"]

    assert cells[0][0]["merged"] is True
    assert cells[0][0]["merge_master"] == "A1"
    assert cells[0][0]["merge_range"] == "A1:B2"
    assert cells[1][1]["merged"] is True
    assert cells[1][1]["merge_master"] == "A1"
    assert cells[1][0]["row_hidden"] is True
    assert cells[1][0]["row_filtered"] is False
    assert cells[4][0]["row_hidden"] is True
    assert cells[4][0]["row_filtered"] is True


def test_non_finite_numeric_cell_does_not_turn_into_json_null(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Malformed"
    sheet["A1"] = 1
    path = tmp_path / "non-finite.xlsx"
    workbook.save(path)
    _rewrite_sheet_xml(path, {b"<v>1</v>": b"<v>1E9999</v>"})

    cell = LocalWorkbookReader().read(path, "Malformed", "A1")[0][0]

    assert cell["value"] == "Infinity"
    assert cell["value_type"] == "non_finite_number"
    assert cell["cached_value"] == "Infinity"
    json.dumps(cell, allow_nan=False)


def test_deep_read_translates_shared_formulas_for_non_master_cells(tmp_path: Path) -> None:
    # Given a follower cell in a shared-formula group (t="shared" with no
    # formula text of its own, only a shared-group index "si" plus its own
    # cached <v>) - the common Excel-generated shape for "fill formula down
    # a column", which stores the formula text once on the master cell only.
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Shared"
    sheet["A1"] = 1
    sheet["A2"] = 2
    sheet["B1"] = "=A1*2"
    sheet["B2"] = "=A2*2"
    path = tmp_path / "shared.xlsx"
    workbook.save(path)
    _rewrite_sheet_xml(
        path,
        {
            b"<c r=\"B1\"><f>A1*2</f><v></v></c>": (
                b'<c r="B1"><f t="shared" ref="B1:B2" si="0">A1*2</f><v>2</v></c>'
            ),
            b"<c r=\"B2\"><f>A2*2</f><v></v></c>": b'<c r="B2"><f t="shared" si="0"/><v>4</v></c>',
        },
    )

    cells = read_xlsx_range(path, "Shared", "A1:B2")["cells"]

    # Then the follower cell's relative reference is translated ("A1*2" ->
    # "A2*2") instead of being left blank/untranslated, and its cached value
    # is still readable.
    assert cells[0][1]["formula"] == "=A1*2"
    assert cells[0][1]["cached_value"] == 2
    assert cells[1][1]["formula"] == "=A2*2"
    assert cells[1][1]["cached_value"] == 4


def test_deep_read_reports_formula_and_literal_error_cells(tmp_path: Path) -> None:
    # Given a formula cell that evaluates to #DIV/0! and a literal #REF!
    # error cell (both use data_type "e" with the error code as the cached
    # text, per ECMA-376 18.3.1.4).
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Errors"
    sheet["A1"] = "=1/0"
    sheet["B1"] = 0
    path = tmp_path / "errors.xlsx"
    workbook.save(path)
    _rewrite_sheet_xml(
        path,
        {
            b"<c r=\"A1\"><f>1/0</f><v></v></c>": b'<c r="A1" t="e"><f>1/0</f><v>#DIV/0!</v></c>',
            b'<c r="B1" t="n"><v>0</v></c>': b'<c r="B1" t="e"><v>#REF!</v></c>',
        },
    )

    cells = read_xlsx_range(path, "Errors", "A1:B1")["cells"][0]

    # Then both cells surface as readable error values instead of raising or
    # silently coming back as None/0.
    assert cells[0]["cached_value"] == "#DIV/0!"
    assert cells[0]["cached_value_type"] == "error"
    assert cells[1]["value"] == "#REF!"
    assert cells[1]["value_type"] == "error"


def test_deep_read_honors_the_1904_date_system_workbook_flag(tmp_path: Path) -> None:
    # Given two otherwise-identical workbooks whose single date cell stores
    # the exact same Excel serial number, differing only in
    # <workbookPr date1904="1"/> (the 1904 date system some workbooks -
    # notably ones with a Mac Excel history - declare). The same serial
    # number means a different real-world date depending on which epoch is
    # in effect; silently ignoring the flag would shift every date in the
    # workbook by (incorrectly) about four years.
    epoch_1900_path = tmp_path / "epoch_1900.xlsx"
    epoch_1904_path = tmp_path / "epoch_1904.xlsx"
    for path in (epoch_1900_path, epoch_1904_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Dates"
        sheet["A1"] = date(2024, 1, 1)
        workbook.save(path)
    _rewrite_member(
        epoch_1904_path, "xl/workbook.xml", {b"<workbookPr/>": b'<workbookPr date1904="1"/>'}
    )

    normal_epoch = read_xlsx_range(epoch_1900_path, "Dates", "A1")["cells"][0][0]
    shifted_epoch = read_xlsx_range(epoch_1904_path, "Dates", "A1")["cells"][0][0]

    # Then the identical stored serial number resolves to a different
    # calendar date under the 1904 flag instead of both files reporting the
    # same 2024-01-01 date.
    assert normal_epoch["excel_serial"] == shifted_epoch["excel_serial"]
    assert normal_epoch["value"] == "2024-01-01T00:00:00"
    assert shifted_epoch["value"] != normal_epoch["value"]


def test_rest_xlsx_range_returns_date_cells_in_the_json_envelope(test_container) -> None:
    path = test_container.settings.project_root / "source" / "schedule.xlsx"
    workbook = Workbook()
    workbook.iso_dates = True
    sheet = workbook.active
    sheet.title = "Schedule"
    sheet.append(["milestone", date(2026, 8, 13)])
    workbook.save(path)
    context = test_container.application.operations.request_context()
    test_container.application.ingestion.sync_filesystem(context, "fixture")
    hit = test_container.application.retrieval.search(
        context,
        SearchRequest(query="milestone"),
    )[0]
    client = TestClient(create_app(test_container))

    response = client.get(
        f"/v1/xlsx/{hit.artifact_id}/range",
        headers={"X-KIP-API-Key": "test-key"},
        params={"sheet": "Schedule", "cell_range": "A1:B1"},
    )

    assert response.status_code == 200
    cell = response.json()["data"]["cells"][0][1]
    assert cell["value"] == "2026-08-13"
    assert cell["cached_value"] == "2026-08-13"
    assert cell["excel_serial"] == pytest.approx(46247.0)


def test_cli_xlsx_range_returns_date_cells_in_the_versioned_envelope(
    monkeypatch,
    test_container,
) -> None:
    path = test_container.settings.project_root / "source" / "schedule-cli.xlsx"
    workbook = Workbook()
    workbook.iso_dates = True
    sheet = workbook.active
    sheet.title = "Schedule"
    sheet.append(["deadline", date(2026, 8, 13)])
    workbook.save(path)
    context = test_container.application.operations.request_context()
    test_container.application.ingestion.sync_filesystem(context, "fixture")
    hit = test_container.application.retrieval.search(
        context,
        SearchRequest(query="deadline"),
    )[0]
    monkeypatch.setattr(
        "kip.cli.build_container",
        lambda settings, load_models=True: test_container,
    )

    result = CliRunner().invoke(
        app,
        ["xlsx-read", hit.artifact_id, "--sheet", "Schedule", "--range", "A1:B1"],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert payload["schema_version"] == "kip.envelope.v1"
    assert payload["data"]["cells"][0][1]["value"] == "2026-08-13"
