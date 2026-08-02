from __future__ import annotations

import re
import zipfile
from collections import OrderedDict
from io import BytesIO
from pathlib import Path, PurePosixPath
from xml.etree import ElementTree as ET

from kip.application.analyzer import normalize_text
from kip.domain.models import ContentUnit, EvidenceLocator, ExtractionRun
from kip.errors import DependencyUnavailableError, ParserError, ValidationError
from kip.ids import new_id, sha256_bytes, stable_id

_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_CELL_RANGE_RE = re.compile(r"^[A-Za-z]{1,3}[1-9][0-9]*(?::[A-Za-z]{1,3}[1-9][0-9]*)?$")


def _tag(local: str) -> str:
    return f"{{{_MAIN_NS}}}{local}"


def _safe_zip_check(archive: zipfile.ZipFile, max_entries: int = 100000, max_uncompressed: int = 2_147_483_648, max_ratio: int = 200) -> None:
    infos = archive.infolist()
    if len(infos) > max_entries:
        raise ValidationError("XLSX has too many ZIP entries")
    total = sum(info.file_size for info in infos)
    compressed = max(1, sum(info.compress_size for info in infos))
    if total > max_uncompressed or total / compressed > max_ratio:
        raise ValidationError("XLSX decompression limits exceeded")


def _read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    name = "xl/sharedStrings.xml"
    if name not in archive.namelist():
        return []
    values: list[str] = []
    with archive.open(name) as stream:
        for event, elem in ET.iterparse(stream, events=("end",)):
            if elem.tag == _tag("si"):
                texts = [node.text or "" for node in elem.iter(_tag("t"))]
                values.append("".join(texts))
                elem.clear()
    return values


def _sheet_paths(archive: zipfile.ZipFile) -> list[tuple[str, str]]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rel_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rels: dict[str, str] = {}
    for rel in rel_root.findall(f"{{{_PACKAGE_REL_NS}}}Relationship"):
        rels[rel.attrib["Id"]] = rel.attrib["Target"]
    result: list[tuple[str, str]] = []
    for sheet in workbook.findall(f".//{{{_MAIN_NS}}}sheet"):
        name = sheet.attrib.get("name", "Sheet")
        rel_id = sheet.attrib.get(f"{{{_REL_NS}}}id")
        if not rel_id or rel_id not in rels:
            continue
        raw_target = rels[rel_id].lstrip("/")
        target = PurePosixPath(raw_target)
        if not raw_target.startswith("xl/"):
            target = PurePosixPath("xl") / target
        normalized = str(PurePosixPath(*[part for part in target.parts if part not in {".", "..", "/"}]))
        result.append((name, normalized))
    return result


def _parse_sheet(archive: zipfile.ZipFile, path: str, shared: list[str], max_chars: int) -> tuple[str | None, list[str], bool]:
    dimension: str | None = None
    values: OrderedDict[str, None] = OrderedDict()
    truncated = False
    current_cell_type: str | None = None
    current_value: str | None = None
    current_inline: list[str] = []
    with archive.open(path) as stream:
        for event, elem in ET.iterparse(stream, events=("start", "end")):
            if event == "start" and elem.tag == _tag("c"):
                current_cell_type = elem.attrib.get("t")
                current_value = None
                current_inline = []
            elif event == "end":
                if elem.tag == _tag("dimension"):
                    dimension = elem.attrib.get("ref")
                elif elem.tag == _tag("v"):
                    current_value = elem.text or ""
                elif elem.tag == _tag("t") and current_cell_type == "inlineStr":
                    current_inline.append(elem.text or "")
                elif elem.tag == _tag("c"):
                    text: str | None = None
                    if current_cell_type == "s" and current_value and current_value.isdigit():
                        idx = int(current_value)
                        if 0 <= idx < len(shared):
                            text = shared[idx]
                    elif current_cell_type in {"inlineStr", "str"}:
                        text = "".join(current_inline) if current_inline else current_value
                    if text:
                        normalized = normalize_text(text)
                        if normalized:
                            values.setdefault(normalized, None)
                            if sum(len(v) + 1 for v in values) > max_chars:
                                truncated = True
                                break
                    current_cell_type = None
                    current_value = None
                    current_inline = []
                elem.clear()
    return dimension, list(values), truncated


class XlsxShallowParser:
    name = "xlsx-shallow"
    version = "1.0"

    def __init__(self, max_chars_per_sheet: int = 240_000) -> None:
        self.max_chars_per_sheet = max_chars_per_sheet

    def supports(self, path: Path) -> bool:
        return path.suffix.lower() == ".xlsx"

    def parse(self, path: Path, *, artifact_id: str, document_id: str, acl_scopes: list[str]) -> tuple[ExtractionRun, list[ContentUnit]]:
        extraction_id = new_id("ext")
        units: list[ContentUnit] = []
        warnings: list[str] = []
        try:
            with zipfile.ZipFile(path) as archive:
                _safe_zip_check(archive)
                shared = _read_shared_strings(archive)
                for ordinal, (sheet_name, sheet_path) in enumerate(_sheet_paths(archive)):
                    dimension, strings, truncated = _parse_sheet(
                        archive, sheet_path, shared, self.max_chars_per_sheet
                    )
                    header = f"Sheet: {sheet_name}\nUsed range: {dimension or 'unknown'}"
                    body = header + ("\n" + "\n".join(strings) if strings else "")
                    normalized = normalize_text(body)
                    if truncated:
                        warnings.append(f"sheet {sheet_name}: shallow text truncated")
                    units.append(
                        ContentUnit(
                            id=stable_id("unit", extraction_id, str(ordinal)),
                            extraction_id=extraction_id,
                            document_id=document_id,
                            artifact_id=artifact_id,
                            ordinal=ordinal,
                            unit_type="xlsx_sheet_shallow",
                            title=f"{path.name} - {sheet_name}",
                            body=body,
                            body_normalized=normalized,
                            lexical_text=normalized,
                            locator=EvidenceLocator(
                                type="xlsx_sheet", data={"sheet": sheet_name, "range": dimension}
                            ),
                            acl_scopes=acl_scopes,
                            metadata={
                                "sheet": sheet_name,
                                "used_range": dimension,
                                "shared_string_count": len(strings),
                                "truncated": truncated,
                                "deep_read_required_for_numbers": True,
                            },
                        )
                    )
        except (zipfile.BadZipFile, KeyError, ET.ParseError) as exc:
            raise ParserError(f"XLSX parse failed: {path}: {exc}") from exc
        aggregate = "\n".join(unit.body for unit in units)
        extraction = ExtractionRun(
            id=extraction_id,
            artifact_id=artifact_id,
            parser_name=self.name,
            parser_version=self.version,
            status="partial" if warnings else "succeeded",
            quality_score=0.9 if units else 0.0,
            output_hash=sha256_bytes(aggregate.encode("utf-8")),
            warnings=warnings,
            metadata={"mode": "shallow", "numeric_values_indexed": False},
        )
        return extraction, units


def read_xlsx_range(path: Path, sheet: str, cell_range: str, *, include_formula_and_cached: bool = True) -> dict:
    if not _CELL_RANGE_RE.fullmatch(cell_range):
        raise ValidationError("invalid XLSX cell range")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DependencyUnavailableError("Install the extractors extra for XLSX deep reads") from exc

    def read(data_only: bool) -> list[list[dict]]:
        workbook = load_workbook(path, read_only=True, data_only=data_only)
        if sheet not in workbook.sheetnames:
            workbook.close()
            raise ValidationError(f"sheet does not exist: {sheet}")
        worksheet = workbook[sheet]
        rows: list[list[dict]] = []
        for row in worksheet[cell_range]:
            rows.append(
                [
                    {
                        "coordinate": cell.coordinate,
                        "value": cell.value,
                        "data_type": cell.data_type,
                    }
                    for cell in row
                ]
            )
        workbook.close()
        return rows

    formulas = read(data_only=False)
    result = {"sheet": sheet, "range": cell_range, "cells": formulas}
    if include_formula_and_cached:
        cached = read(data_only=True)
        for row_index, row in enumerate(result["cells"]):
            for col_index, cell in enumerate(row):
                cell["cached_value"] = cached[row_index][col_index]["value"]
    return result
