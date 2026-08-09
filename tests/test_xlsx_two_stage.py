from __future__ import annotations

import hashlib
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

from kip.adapters.parsers.xlsx import XlsxShallowParser, read_xlsx_range
from kip.domain.models import SearchRequest
from kip.errors import ValidationError


def test_xlsx_shallow_search_and_deep_range(test_container):
    path = test_container.settings.project_root / "source" / "A과제_정산.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "정산"
    sheet.append(["구분", "항목", "금액"])
    sheet.append(["장비비", "카메라", 890000])
    sheet.append(["인건비", "참여율", "=C2*2"])
    workbook.save(path)

    context = test_container.application.operations.request_context()
    summary = test_container.application.ingestion.sync_filesystem(context, "fixture")
    assert summary.inserted == 1

    hits = test_container.application.retrieval.search(context, SearchRequest(query="장비비 카메라", limit=10))
    assert hits
    unit = test_container.repository.get_content_unit(context, hits[0].unit_id)
    assert unit.unit_type == "xlsx_sheet_shallow"
    assert "장비비" in unit.body
    assert "890000" not in unit.body
    assert unit.metadata["deep_read_required_for_numbers"] is True

    deep = test_container.application.evidence.read_xlsx(
        context,
        hits[0].artifact_id,
        sheet="정산",
        cell_range="A1:C3",
    )
    assert deep.cells[1][2]["value"] == 890000
    assert deep.cells[2][2]["value"] == "=C2*2"
    assert deep.source_changed_since_index is False


def test_xlsx_deep_read_preserves_coordinates_for_sparse_range(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sparse"
    sheet["D4"] = "끝 셀"
    path = tmp_path / "sparse.xlsx"
    workbook.save(path)

    result = read_xlsx_range(path, "Sparse", "A1:D4")

    assert result["cells"][0][0]["coordinate"] == "A1"
    assert result["cells"][0][1]["value"] is None
    assert result["cells"][3][3]["coordinate"] == "D4"
    assert result["cells"][3][3]["value"] == "끝 셀"


def test_xlsm_deep_read_preserves_formula_and_cached_value_shape(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MacroSheet"
    sheet["A1"] = "금액"
    sheet["B1"] = "=1+2"
    path = tmp_path / "macro.xlsm"
    workbook.save(path)

    result = read_xlsx_range(path, "MacroSheet", "A1:B1")

    assert result["cells"][0][0]["coordinate"] == "A1"
    assert result["cells"][0][1]["coordinate"] == "B1"
    assert result["cells"][0][1]["value"] == "=1+2"


def test_xlsm_sync_and_service_deep_read(test_container):
    test_container.settings.raw["sources"]["filesystem"][0]["include_extensions"].append(".xlsm")
    path = test_container.settings.project_root / "source" / "macro.xlsm"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MacroSheet"
    sheet["A1"] = "승인"
    sheet["B1"] = "=1+2"
    workbook.save(path)

    context = test_container.application.operations.request_context()
    summary = test_container.application.ingestion.sync_filesystem(context, "fixture")
    assert summary.inserted == 1

    hits = test_container.application.retrieval.search(context, SearchRequest(query="승인", limit=10))
    assert hits
    deep = test_container.application.evidence.read_xlsx(context, hits[0].artifact_id, sheet="MacroSheet", cell_range="A1:B1")

    assert deep.cells[0][1]["value"] == "=1+2"


def test_xlsx_shallow_parser_chunks_large_sheet_units(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Large"
    for row in range(1, 31):
        sheet.cell(row=row, column=1, value=f"고유 증거 항목 {row:03d}")
    path = tmp_path / "large.xlsx"
    workbook.save(path)

    extraction, units = XlsxShallowParser(
        max_chars_per_sheet=10_000,
        max_chars_per_unit=80,
    ).parse(path, artifact_id="art_large", document_id="doc_large", acl_scopes=["workspace:default"])

    assert extraction.status == "succeeded"
    assert len(units) > 1
    assert all(len(unit.body) <= 80 for unit in units)
    assert "고유 증거 항목 001" in "\n".join(unit.body for unit in units)
    assert all(unit.locator.type == "xlsx_sheet" for unit in units)
    assert [unit.ordinal for unit in units] == list(range(len(units)))


def test_xlsx_deep_read_preserves_cell_semantics_and_source_bytes(tmp_path: Path) -> None:
    # Given a workbook with a formatted date and hidden row and column metadata.
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Evidence"
    sheet["B2"] = date(2026, 8, 6)
    sheet["B2"].number_format = "yyyy-mm-dd"
    sheet.row_dimensions[2].hidden = True
    sheet.column_dimensions["B"].hidden = True
    path = tmp_path / "evidence.xlsx"
    workbook.save(path)
    before = hashlib.sha256(path.read_bytes()).hexdigest()

    # When the exact range is read through the read-only deep reader.
    result = read_xlsx_range(path, "Evidence", "B2:B2")

    # Then semantic cell metadata is explicit and the authoritative source is unchanged.
    cell = result["cells"][0][0]
    assert cell["number_format"] == "yyyy-mm-dd"
    assert cell["is_date"] is True
    assert cell["row_hidden"] is True
    assert cell["column_hidden"] is True
    assert hashlib.sha256(path.read_bytes()).hexdigest() == before


def test_xlsx_deep_read_rejects_archive_over_expansion_limit(tmp_path: Path) -> None:
    # Given a highly compressible workbook-shaped ZIP that exceeds the safety ratio.
    path = tmp_path / "oversized.xlsx"
    import zipfile

    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("xl/workbook.xml", b"0" * 1_000_000)

    # When the deep reader inspects it.
    # Then archive safety is enforced before openpyxl expands the payload.
    with pytest.raises(ValidationError, match="decompression limits"):
        read_xlsx_range(path, "Sheet1", "A1:A1")
